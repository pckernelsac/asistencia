# -*- coding: utf-8 -*-
from __future__ import annotations

import logging
import os
import re
import sqlite3
import time
import unicodedata
from datetime import datetime
from functools import wraps
from io import BytesIO
from typing import Any, Optional

import click
import openpyxl
import qrcode
from dotenv import load_dotenv
from flask import (
    Flask,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    Response,
    send_from_directory,
    session,
    url_for,
)
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from zoneinfo import ZoneInfo

import db as dbx
import storage

load_dotenv()

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB_PATH = os.path.join(BASE_DIR, "instance", "asistencia.db")
IS_VERCEL = bool(os.environ.get("VERCEL"))
UPLOAD_FOLDER = os.path.join("/tmp", "uploads") if IS_VERCEL else os.path.join(BASE_DIR, "static", "uploads")
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp", "gif"}

if not IS_VERCEL:
    os.makedirs(os.path.join(BASE_DIR, "instance"), exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

dbx.set_sqlite_path(DB_PATH)


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def slugify_institution(name: str) -> str:
    s = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s or "institucion"


def retry_on_db_locked(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        max_retries = 3
        for attempt in range(max_retries):
            try:
                return func(*args, **kwargs)
            except Exception as e:
                msg = str(e).lower()
                if attempt < max_retries - 1 and ("locked" in msg or "deadlock" in msg):
                    time.sleep(0.1 * (attempt + 1))
                    continue
                raise
        return func(*args, **kwargs)

    return wrapper


def _init_sqlite_tables(conn) -> None:
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS institutions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            slug TEXT NOT NULL UNIQUE,
            approval_status TEXT NOT NULL DEFAULT 'pending',
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            is_super_master INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS institution_members (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            institution_id INTEGER NOT NULL,
            role TEXT NOT NULL DEFAULT 'admin',
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, institution_id),
            FOREIGN KEY(user_id) REFERENCES users(id),
            FOREIGN KEY(institution_id) REFERENCES institutions(id)
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombres TEXT NOT NULL,
            apellidos TEXT NOT NULL,
            dni TEXT NOT NULL UNIQUE,
            photo TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL,
            recorded_at TEXT NOT NULL,
            tipo TEXT NOT NULL DEFAULT 'ENTRADA',
            FOREIGN KEY(student_id) REFERENCES students(id)
        );
        """
    )
    dbx.commit(conn)


def _sqlite_table_columns(cur, table: str) -> list:
    cur.execute("PRAGMA table_info(" + table + ")")
    return [r[1] for r in cur.fetchall()]


def _migrate_sqlite_institution_approval(conn) -> None:
    cur = conn.cursor()
    cols = _sqlite_table_columns(cur, "institutions")
    if "approval_status" in cols:
        return
    cur.execute(
        "ALTER TABLE institutions ADD COLUMN approval_status TEXT NOT NULL DEFAULT 'pending'"
    )
    dbx.execute(cur, "UPDATE institutions SET approval_status = %s", ("approved",))
    dbx.commit(conn)


def _migrate_sqlite_user_super_master(conn) -> None:
    cur = conn.cursor()
    cols = _sqlite_table_columns(cur, "users")
    if "is_super_master" in cols:
        return
    cur.execute("ALTER TABLE users ADD COLUMN is_super_master INTEGER NOT NULL DEFAULT 0")
    dbx.commit(conn)
    dbx.execute(cur, "UPDATE users SET is_super_master = 1 WHERE username = %s", ("admin",))
    for un in [x.strip() for x in (os.environ.get("SUPER_MASTER_USERNAMES") or "").split(",") if x.strip()]:
        dbx.execute(
            cur,
            "UPDATE users SET is_super_master = 1 WHERE lower(username) = lower(%s)",
            (un,),
        )
    dbx.commit(conn)


def _migrate_sqlite_saas_columns(conn) -> None:
    cur = conn.cursor()

    def col_names(table: str) -> list:
        if table not in ("users", "students", "attendance"):
            return []
        cur.execute("PRAGMA table_info(" + table + ")")
        return [r[1] for r in cur.fetchall()]

    if "institution_id" not in col_names("users"):
        cur.execute("ALTER TABLE users ADD COLUMN institution_id INTEGER REFERENCES institutions(id)")
    if "institution_id" not in col_names("students"):
        cur.execute("ALTER TABLE students ADD COLUMN institution_id INTEGER REFERENCES institutions(id)")
    if "institution_id" not in col_names("attendance"):
        cur.execute("ALTER TABLE attendance ADD COLUMN institution_id INTEGER REFERENCES institutions(id)")

    dbx.commit(conn)


def _migrate_sqlite_tipo(conn) -> None:
    cur = conn.cursor()
    cur.execute("PRAGMA table_info(attendance)")
    columns = [row[1] for row in cur.fetchall()]
    if "tipo" not in columns:
        try:
            cur.execute("ALTER TABLE attendance ADD COLUMN tipo TEXT NOT NULL DEFAULT 'ENTRADA'")
            dbx.commit(conn)
        except sqlite3.OperationalError:
            pass


def _ensure_sqlite_default_institution(conn) -> int:
    cur = conn.cursor()
    dbx.execute(cur, "SELECT id FROM institutions ORDER BY id LIMIT 1")
    row = dbx.fetchone(cur)
    if row:
        return row["id"]
    dbx.execute(
        cur,
        "INSERT INTO institutions (name, slug, approval_status) VALUES (%s, %s, %s)",
        ("Mi institución", "mi-institucion", "approved"),
    )
    dbx.commit(conn)
    dbx.execute(cur, "SELECT id FROM institutions ORDER BY id DESC LIMIT 1")
    return dbx.fetchone(cur)["id"]


def _backfill_sqlite_institution_ids(conn, default_iid: int) -> None:
    cur = conn.cursor()
    dbx.execute(cur, "UPDATE students SET institution_id = %s WHERE institution_id IS NULL", (default_iid,))
    dbx.execute(
        cur,
        """
        UPDATE attendance SET institution_id = (
            SELECT s.institution_id FROM students s WHERE s.id = attendance.student_id
        )
        WHERE institution_id IS NULL
        """,
    )
    dbx.execute(cur, "UPDATE users SET institution_id = %s WHERE institution_id IS NULL", (default_iid,))
    dbx.commit(conn)


def _ensure_sqlite_admin_membership(conn, default_iid: int) -> None:
    cur = conn.cursor()
    dbx.execute(cur, "SELECT id FROM users WHERE username = %s", ("admin",))
    u = dbx.fetchone(cur)
    if not u:
        return
    uid = u["id"]
    dbx.execute(
        cur,
        "SELECT 1 FROM institution_members WHERE user_id = %s AND institution_id = %s",
        (uid, default_iid),
    )
    if dbx.fetchone(cur):
        return
    dbx.execute(
        cur,
        "INSERT INTO institution_members (user_id, institution_id, role) VALUES (%s, %s, %s)",
        (uid, default_iid, "owner"),
    )
    dbx.commit(conn)


def _create_indexes_sqlite(conn) -> None:
    cur = conn.cursor()
    stmts = [
        "CREATE INDEX IF NOT EXISTS idx_students_institution ON students(institution_id)",
        "CREATE INDEX IF NOT EXISTS idx_students_inst_dni ON students(institution_id, dni)",
        "CREATE INDEX IF NOT EXISTS idx_attendance_institution ON attendance(institution_id)",
        "CREATE INDEX IF NOT EXISTS idx_attendance_student ON attendance(student_id)",
        "CREATE INDEX IF NOT EXISTS idx_attendance_date ON attendance(recorded_at)",
        "CREATE INDEX IF NOT EXISTS idx_attendance_student_tipo ON attendance(student_id, tipo)",
        "CREATE INDEX IF NOT EXISTS idx_institution_members_user ON institution_members(user_id)",
    ]
    for s in stmts:
        try:
            cur.execute(dbx.q(s))
        except sqlite3.OperationalError:
            pass
    dbx.commit(conn)


def init_db() -> None:
    if dbx.is_postgres():
        return
    with dbx.get_connection() as conn:
        _init_sqlite_tables(conn)
        _migrate_sqlite_tipo(conn)
        _migrate_sqlite_saas_columns(conn)
        _migrate_sqlite_institution_approval(conn)
        _migrate_sqlite_user_super_master(conn)
        iid = _ensure_sqlite_default_institution(conn)
        _backfill_sqlite_institution_ids(conn, iid)
        _create_indexes_sqlite(conn)


def create_default_user() -> None:
    if dbx.is_postgres():
        return
    with dbx.get_connection() as conn:
        cur = conn.cursor()
        dbx.execute(cur, "SELECT id FROM users WHERE username = %s", ("admin",))
        if dbx.fetchone(cur):
            return
        password_hash = generate_password_hash("admin123")
        dbx.execute(
            cur,
            "INSERT INTO users (username, password_hash, is_super_master) VALUES (%s, %s, %s)",
            ("admin", password_hash, 1),
        )
        dbx.commit(conn)
        iid = _ensure_sqlite_default_institution(conn)
        _ensure_sqlite_admin_membership(conn, iid)


def _apply_postgres_schema_migration_body(conn) -> None:
    cur = conn.cursor()
    cur.execute(
        "ALTER TABLE institutions ADD COLUMN IF NOT EXISTS approval_status TEXT NOT NULL DEFAULT 'approved'"
    )
    cur.execute("ALTER TABLE institutions ALTER COLUMN approval_status SET DEFAULT 'pending'")
    cur.execute(
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS is_super_master BOOLEAN NOT NULL DEFAULT FALSE"
    )
    dbx.execute(
        cur,
        "UPDATE users SET is_super_master = TRUE WHERE lower(username) = lower(%s) AND is_super_master = FALSE",
        ("admin",),
    )
    for un in [x.strip() for x in (os.environ.get("SUPER_MASTER_USERNAMES") or "").split(",") if x.strip()]:
        dbx.execute(
            cur,
            "UPDATE users SET is_super_master = TRUE WHERE lower(username) = lower(%s) AND is_super_master = FALSE",
            (un,),
        )
    dbx.commit(conn)


def ensure_postgres_schema_migrations() -> None:
    """Añade columnas nuevas en Postgres si faltan. Usa DIRECT_URL si existe (DDL y pooler no suelen ir bien)."""
    if not dbx.is_postgres():
        return
    direct = (os.environ.get("DIRECT_URL") or "").strip()
    try:
        if direct:
            import psycopg
            from psycopg.rows import dict_row

            with psycopg.connect(direct, row_factory=dict_row) as conn:
                _apply_postgres_schema_migration_body(conn)
        else:
            with dbx.get_connection() as conn:
                _apply_postgres_schema_migration_body(conn)
    except Exception as e:
        logging.getLogger(__name__).warning("ensure_postgres_schema_migrations: %s", e, exc_info=True)


app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-key-2024")
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024

app.config["SUPABASE_URL"] = (os.environ.get("SUPABASE_URL") or "").rstrip("/")
app.config["SUPABASE_ANON_KEY"] = os.environ.get("SUPABASE_ANON_KEY", "")
app.config["SUPABASE_SERVICE_ROLE_KEY"] = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
app.config["DATABASE_URL"] = os.environ.get("DATABASE_URL", "")
app.config["DIRECT_URL"] = os.environ.get("DIRECT_URL", "")

if os.environ.get("FLASK_ENV") == "production":
    app.config["SESSION_COOKIE_SECURE"] = True
    app.config["SESSION_COOKIE_HTTPONLY"] = True
    app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
    app.config["PERMANENT_SESSION_LIFETIME"] = 3600

init_db()
create_default_user()
ensure_postgres_schema_migrations()


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            flash("Debe iniciar sesión para acceder", "error")
            return redirect(url_for("login"))
        return f(*args, **kwargs)

    return decorated_function


def super_master_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            flash("Debe iniciar sesión para acceder", "error")
            return redirect(url_for("login"))
        if not session.get("is_super_master"):
            flash("No tiene permisos de super administrador.", "error")
            return redirect(url_for("select_institution"))
        return f(*args, **kwargs)

    return decorated_function


def tenant_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            flash("Debe iniciar sesión para acceder", "error")
            return redirect(url_for("login"))
        if not session.get("institution_id"):
            flash("Seleccione o cree una institución para continuar.", "error")
            return redirect(url_for("select_institution"))
        iid = int(session["institution_id"])
        if not institution_is_approved(iid):
            flash("Esta institución no está aprobada o fue desactivada.", "error")
            session.pop("institution_id", None)
            session.pop("institution_name", None)
            return redirect(url_for("select_institution"))
        return f(*args, **kwargs)

    return decorated_function


@app.context_processor
def inject_tenant():
    home = url_for("login")
    if session.get("user_id"):
        home = url_for("select_institution")
        if session.get("institution_id"):
            home = url_for("dashboard")
    return {
        "current_institution_name": session.get("institution_name", ""),
        "current_institution_id": session.get("institution_id"),
        "nav_home_url": home,
        "is_super_master": bool(session.get("is_super_master")),
        "photo_url": storage.public_url,
    }


def _load_peru_tz():
    try:
        return ZoneInfo("America/Lima")
    except Exception:
        try:
            import tzdata  # noqa: F401

            return ZoneInfo("America/Lima")
        except Exception:
            from datetime import timedelta, timezone

            return timezone(timedelta(hours=-5), name="America/Lima")


TZ = _load_peru_tz()


def now_pe() -> datetime:
    return datetime.now(TZ)


def current_institution_id() -> int:
    iid = session.get("institution_id")
    if not iid:
        raise RuntimeError("Sin institución en sesión")
    return int(iid)


def row_approval_status(m: Any) -> str:
    try:
        v = m["approval_status"]
        return str(v) if v is not None else "approved"
    except (KeyError, TypeError):
        return "approved"


def user_is_super_master(user_row: Any) -> bool:
    try:
        v = user_row["is_super_master"]
        if isinstance(v, bool):
            return v
        return int(v) != 0
    except (KeyError, TypeError, ValueError):
        pass
    try:
        un = str(user_row["username"]).lower()
    except (KeyError, TypeError):
        return False
    names = (os.environ.get("SUPER_MASTER_USERNAMES") or "").strip()
    if not names:
        return False
    allowed = {x.strip().lower() for x in names.split(",") if x.strip()}
    return un in allowed


def institution_is_approved(institution_id: int) -> bool:
    with dbx.get_connection() as conn:
        cur = conn.cursor()
        dbx.execute(
            cur,
            "SELECT approval_status FROM institutions WHERE id = %s",
            (institution_id,),
        )
        row = dbx.fetchone(cur)
    if not row:
        return False
    try:
        st = row["approval_status"]
    except (KeyError, TypeError):
        return True
    return (st or "approved") == "approved"


def load_user_memberships(user_id: int) -> list:
    with dbx.get_connection() as conn:
        cur = conn.cursor()
        dbx.execute(
            cur,
            """
            SELECT m.institution_id, i.name, i.slug, m.role, i.approval_status
            FROM institution_members m
            JOIN institutions i ON i.id = m.institution_id
            WHERE m.user_id = %s
            ORDER BY i.name
            """,
            (user_id,),
        )
        return dbx.fetchall(cur)


def list_pending_institutions_admin() -> list:
    with dbx.get_connection() as conn:
        cur = conn.cursor()
        dbx.execute(
            cur,
            """
            SELECT i.id, i.name, i.slug, i.created_at, MIN(u.username) AS owner_username
            FROM institutions i
            JOIN institution_members m ON m.institution_id = i.id AND m.role = 'owner'
            JOIN users u ON u.id = m.user_id
            WHERE i.approval_status = 'pending'
            GROUP BY i.id, i.name, i.slug, i.created_at
            ORDER BY i.id ASC
            """,
        )
        return dbx.fetchall(cur)


def set_session_institution(institution_id: int, name: str) -> None:
    session["institution_id"] = institution_id
    session["institution_name"] = name


def find_student_by_dni(dni: str, institution_id: int) -> Optional[Any]:
    with dbx.get_connection() as conn:
        cur = conn.cursor()
        dbx.execute(
            cur,
            "SELECT * FROM students WHERE dni = %s AND institution_id = %s",
            (dni, institution_id),
        )
        return dbx.fetchone(cur)


@retry_on_db_locked
def insert_attendance(student_id: int, institution_id: int, tipo: str) -> int:
    with dbx.get_connection() as conn:
        cur = conn.cursor()
        ts = now_pe()
        if dbx.is_postgres():
            dbx.execute(
                cur,
                """
                INSERT INTO attendance (student_id, institution_id, recorded_at, tipo)
                VALUES (%s, %s, %s, %s)
                RETURNING id
                """,
                (student_id, institution_id, ts, tipo),
            )
            row = dbx.fetchone(cur)
            dbx.commit(conn)
            return int(row["id"])
        dbx.execute(
            cur,
            """
            INSERT INTO attendance (student_id, institution_id, recorded_at, tipo)
            VALUES (%s, %s, %s, %s)
            """,
            (student_id, institution_id, ts.isoformat(), tipo),
        )
        dbx.commit(conn)
        return int(cur.lastrowid)


def check_attendance_today(student_id: int, institution_id: int, tipo: str) -> bool:
    today = now_pe().date().isoformat()
    with dbx.get_connection() as conn:
        cur = conn.cursor()
        if dbx.is_postgres():
            dbx.execute(
                cur,
                """
                SELECT COUNT(*)::int AS count FROM attendance
                WHERE student_id = %s AND institution_id = %s AND tipo = %s
                  AND (recorded_at AT TIME ZONE 'America/Lima')::date = %s::date
                """,
                (student_id, institution_id, tipo, today),
            )
        else:
            dbx.execute(
                cur,
                """
                SELECT COUNT(*) as count FROM attendance
                WHERE student_id = %s AND institution_id = %s AND tipo = %s AND date(recorded_at) = %s
                """,
                (student_id, institution_id, tipo, today),
            )
        r = dbx.fetchone(cur)
        return int(r["count"]) > 0


def list_recent_attendance(institution_id: int, limit: int = 100, q: Optional[str] = None):
    with dbx.get_connection() as conn:
        cur = conn.cursor()
        base_sql = """
            SELECT a.id, a.recorded_at, a.tipo, s.nombres, s.apellidos, s.dni, s.photo
            FROM attendance a
            JOIN students s ON s.id = a.student_id
            WHERE a.institution_id = %s
        """
        params: list = [institution_id]
        if q:
            base_sql += " AND (s.nombres LIKE %s OR s.apellidos LIKE %s OR s.dni LIKE %s)"
            like = f"%{q}%"
            params.extend([like, like, like])
        base_sql += " ORDER BY a.id DESC LIMIT %s"
        params.append(limit)
        dbx.execute(cur, base_sql, params)
        return dbx.fetchall(cur)


def photo_name_for_save(institution_id: int, dni: str, filename: str) -> str:
    safe = secure_filename(filename)
    return f"{institution_id}_{dni}_{safe}"


@app.route("/login", methods=["GET", "POST"])
def login():
    if "user_id" in session:
        with dbx.get_connection() as conn:
            cur = conn.cursor()
            dbx.execute(cur, "SELECT * FROM users WHERE id = %s", (int(session["user_id"]),))
            urow = dbx.fetchone(cur)
        if urow:
            session["is_super_master"] = user_is_super_master(urow)
        if session.get("institution_id"):
            return redirect(url_for("dashboard"))
        return redirect(url_for("select_institution"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        if not username or not password:
            flash("Usuario y contraseña son requeridos", "error")
            return redirect(url_for("login"))

        with dbx.get_connection() as conn:
            cur = conn.cursor()
            dbx.execute(cur, "SELECT * FROM users WHERE username = %s", (username,))
            user = dbx.fetchone(cur)

        if not user or not check_password_hash(user["password_hash"], password):
            flash("Usuario o contraseña incorrectos", "error")
            return redirect(url_for("login"))

        session["user_id"] = user["id"]
        session["username"] = user["username"]
        session["is_super_master"] = user_is_super_master(user)
        session.pop("institution_id", None)
        session.pop("institution_name", None)

        members = load_user_memberships(user["id"])
        approved = [m for m in members if row_approval_status(m) == "approved"]
        pending = [m for m in members if row_approval_status(m) == "pending"]
        rejected = [m for m in members if row_approval_status(m) == "rejected"]

        if not members:
            flash("Bienvenido. Cree o solicite acceso a una institución para continuar.", "success")
            return redirect(url_for("register_institution"))

        if len(approved) == 1:
            m = approved[0]
            set_session_institution(int(m["institution_id"]), m["name"])
            flash(f"Bienvenido {username}!", "success")
            return redirect(url_for("dashboard"))

        if len(approved) > 1:
            flash(f"Bienvenido {username}. Elija una institución.", "success")
            return redirect(url_for("select_institution"))

        if len(approved) == 0:
            if session.get("is_super_master"):
                flash(
                    "Puede aprobar instituciones desde Administración. Si tiene una solicitud propia pendiente, espere o pídala a otro super administrador.",
                    "success",
                )
                return redirect(url_for("admin_pending_institutions"))
            if pending:
                flash(
                    "Tiene instituciones pendientes de aprobación. Un super administrador debe aprobarlas antes de usar el panel.",
                    "error",
                )
                return redirect(url_for("select_institution"))
            if rejected:
                flash(
                    "No tiene instituciones aprobadas. Alguna solicitud pudo haber sido rechazada; puede crear una nueva.",
                    "error",
                )
                return redirect(url_for("select_institution"))
            flash("No tiene instituciones aprobadas. Contacte al administrador.", "error")
            return redirect(url_for("select_institution"))

    return render_template("login.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    if "user_id" in session:
        return redirect(url_for("select_institution"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        password2 = request.form.get("password2", "").strip()

        if not username or not password:
            flash("Usuario y contraseña son requeridos", "error")
            return redirect(url_for("register"))

        if password != password2:
            flash("Las contraseñas no coinciden", "error")
            return redirect(url_for("register"))

        if len(password) < 6:
            flash("La contraseña debe tener al menos 6 caracteres", "error")
            return redirect(url_for("register"))

        with dbx.get_connection() as conn:
            cur = conn.cursor()
            dbx.execute(cur, "SELECT id FROM users WHERE username = %s", (username,))
            if dbx.fetchone(cur):
                flash("El nombre de usuario ya existe", "error")
                return redirect(url_for("register"))

            password_hash = generate_password_hash(password)
            dbx.execute(
                cur,
                "INSERT INTO users (username, password_hash, is_super_master) VALUES (%s, %s, %s)",
                (username, password_hash, False),
            )
            dbx.commit(conn)
            dbx.execute(cur, "SELECT * FROM users WHERE username = %s", (username,))
            user = dbx.fetchone(cur)

        session["user_id"] = user["id"]
        session["username"] = user["username"]
        session["is_super_master"] = False
        flash("Cuenta creada. Ahora registre su institución para comenzar.", "success")
        return redirect(url_for("register_institution"))

    return render_template("register.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("Sesión cerrada correctamente", "success")
    return redirect(url_for("login"))


@app.route("/select-institution", methods=["GET", "POST"])
@login_required
def select_institution():
    uid = int(session["user_id"])
    members = load_user_memberships(uid)
    approved = [m for m in members if row_approval_status(m) == "approved"]
    pending_members = [m for m in members if row_approval_status(m) == "pending"]
    rejected_members = [m for m in members if row_approval_status(m) == "rejected"]

    if not members:
        return redirect(url_for("register_institution"))

    if request.method == "POST":
        raw = request.form.get("institution_id", "").strip()
        try:
            choice = int(raw)
        except ValueError:
            flash("Selección inválida", "error")
            return redirect(url_for("select_institution"))
        for m in approved:
            if int(m["institution_id"]) == choice:
                set_session_institution(choice, m["name"])
                return redirect(url_for("dashboard"))
        flash("No pertenece a esa institución o no está aprobada.", "error")
        return redirect(url_for("select_institution"))

    if len(approved) == 1:
        m = approved[0]
        set_session_institution(int(m["institution_id"]), m["name"])
        return redirect(url_for("dashboard"))

    return render_template(
        "select_institution.html",
        approved_members=approved,
        pending_members=pending_members,
        rejected_members=rejected_members,
    )


@app.route("/register-institution", methods=["GET", "POST"])
@login_required
def register_institution():
    uid = int(session["user_id"])
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("El nombre de la institución es obligatorio", "error")
            return redirect(url_for("register_institution"))
        base_slug = slugify_institution(name)
        slug = base_slug
        try:
            with dbx.get_connection() as conn:
                cur = conn.cursor()
                for n in range(20):
                    dbx.execute(cur, "SELECT id FROM institutions WHERE slug = %s", (slug,))
                    if dbx.fetchone(cur):
                        slug = f"{base_slug}-{n+2}"
                        continue
                    dbx.execute(
                        cur,
                        "INSERT INTO institutions (name, slug, approval_status) VALUES (%s, %s, %s)",
                        (name, slug, "pending"),
                    )
                    dbx.commit(conn)
                    dbx.execute(cur, "SELECT id FROM institutions WHERE slug = %s", (slug,))
                    row = dbx.fetchone(cur)
                    iid = int(row["id"])
                    dbx.execute(
                        cur,
                        """
                        INSERT INTO institution_members (user_id, institution_id, role)
                        VALUES (%s, %s, %s)
                        """,
                        (uid, iid, "owner"),
                    )
                    dbx.commit(conn)
                    break
                else:
                    flash("No se pudo generar un identificador único. Pruebe otro nombre.", "error")
                    return redirect(url_for("register_institution"))
            flash(
                "Solicitud enviada. Su institución quedará activa cuando un super administrador la apruebe.",
                "success",
            )
            return redirect(url_for("select_institution"))
        except Exception as e:
            if dbx.is_unique_violation(e):
                flash("Ese identificador ya existe. Pruebe otro nombre.", "error")
            else:
                flash("Error al crear la institución.", "error")
                app.logger.exception(e)
            return redirect(url_for("register_institution"))

    return render_template("register_institution.html")


@app.route("/admin/instituciones-pendientes", methods=["GET", "POST"])
@super_master_required
def admin_pending_institutions():
    if request.method == "POST":
        action = (request.form.get("action") or "").strip()
        raw_id = request.form.get("institution_id", "").strip()
        try:
            iid = int(raw_id)
        except ValueError:
            flash("Identificador inválido.", "error")
            return redirect(url_for("admin_pending_institutions"))
        if action == "approve":
            with dbx.get_connection() as conn:
                cur = conn.cursor()
                dbx.execute(
                    cur,
                    """
                    UPDATE institutions SET approval_status = %s
                    WHERE id = %s AND approval_status = %s
                    """,
                    ("approved", iid, "pending"),
                )
                dbx.commit(conn)
            flash("Institución aprobada. El solicitante ya puede usar el panel.", "success")
        elif action == "reject":
            with dbx.get_connection() as conn:
                cur = conn.cursor()
                dbx.execute(
                    cur,
                    """
                    UPDATE institutions SET approval_status = %s
                    WHERE id = %s AND approval_status = %s
                    """,
                    ("rejected", iid, "pending"),
                )
                dbx.commit(conn)
            flash("Solicitud rechazada.", "success")
        else:
            flash("Acción no reconocida.", "error")
        return redirect(url_for("admin_pending_institutions"))

    pending = list_pending_institutions_admin()
    return render_template("admin_pending_institutions.html", pending=pending)


@app.route("/")
@tenant_required
def dashboard():
    iid = current_institution_id()
    q = request.args.get("q") or ""
    rows = list_recent_attendance(iid, limit=100, q=q if q.strip() else None)

    with dbx.get_connection() as conn:
        cur = conn.cursor()
        dbx.execute(cur, "SELECT COUNT(*) as count FROM students WHERE institution_id = %s", (iid,))
        total_students = dbx.fetchone(cur)["count"]
        dbx.execute(cur, "SELECT COUNT(*) as count FROM attendance WHERE institution_id = %s", (iid,))
        total_attendance = dbx.fetchone(cur)["count"]
        dbx.execute(
            cur,
            """
            SELECT tipo, COUNT(*) as count
            FROM attendance
            WHERE institution_id = %s
            GROUP BY tipo
            """,
            (iid,),
        )
        attendance_by_type = dbx.fetchall(cur)
        today = now_pe().date()
        if dbx.is_postgres():
            dbx.execute(
                cur,
                """
                SELECT COUNT(*)::int AS count
                FROM attendance
                WHERE institution_id = %s
                  AND (recorded_at AT TIME ZONE 'America/Lima')::date = %s::date
                """,
                (iid, str(today)),
            )
        else:
            dbx.execute(
                cur,
                """
                SELECT COUNT(*) as count FROM attendance
                WHERE institution_id = %s AND date(recorded_at) = %s
                """,
                (iid, str(today)),
            )
        today_attendance = dbx.fetchone(cur)["count"]

        if dbx.is_postgres():
            dbx.execute(
                cur,
                """
                SELECT (recorded_at AT TIME ZONE 'America/Lima')::date AS fecha, COUNT(*)::int AS count
                FROM attendance
                WHERE institution_id = %s
                  AND recorded_at >= NOW() AT TIME ZONE 'America/Lima' - INTERVAL '7 days'
                GROUP BY (recorded_at AT TIME ZONE 'America/Lima')::date
                ORDER BY fecha ASC
                """,
                (iid,),
            )
        else:
            dbx.execute(
                cur,
                """
                SELECT date(recorded_at) as fecha, COUNT(*) as count
                FROM attendance
                WHERE institution_id = %s AND date(recorded_at) >= date('now', '-7 days')
                GROUP BY date(recorded_at)
                ORDER BY fecha ASC
                """,
                (iid,),
            )
        attendance_by_day = dbx.fetchall(cur)

        dbx.execute(
            cur,
            """
            SELECT s.nombres, s.apellidos, COUNT(a.id) as total
            FROM students s
            JOIN attendance a ON s.id = a.student_id AND a.institution_id = %s
            WHERE s.institution_id = %s
            GROUP BY s.id, s.nombres, s.apellidos
            ORDER BY total DESC
            LIMIT 5
            """,
            (iid, iid),
        )
        top_students = dbx.fetchall(cur)

    stats = {
        "total_students": total_students,
        "total_attendance": total_attendance,
        "today_attendance": today_attendance,
        "attendance_by_type": [{"tipo": r["tipo"], "count": r["count"]} for r in attendance_by_type],
        "attendance_by_day": [{"fecha": str(r["fecha"]), "count": r["count"]} for r in attendance_by_day],
        "top_students": [{"nombre": f"{r['nombres']} {r['apellidos']}", "total": r["total"]} for r in top_students],
    }

    return render_template("dashboard.html", rows=rows, q=q, stats=stats)


@app.route("/scan")
@tenant_required
def scan():
    return render_template("scan.html")


@app.post("/api/scan")
@tenant_required
def api_scan():
    iid = current_institution_id()
    payload = request.get_json(silent=True) or {}
    code = (payload.get("code") or "").strip()
    if not code:
        return jsonify({"ok": False, "message": "Código vacío."}), 400

    student = find_student_by_dni(code, iid)
    if not student:
        return jsonify({"ok": False, "message": f"No existe estudiante con DNI {code}."}), 404

    now = now_pe()
    hora_actual = now.hour
    tipo = "ENTRADA" if hora_actual < 13 else "SALIDA"

    if check_attendance_today(student["id"], iid, tipo):
        tipo_texto = "entrada" if tipo == "ENTRADA" else "salida"
        ts = now.strftime("%Y-%m-%d %H:%M:%S")
        photo_url = None
        if student["photo"]:
            photo_url = storage.public_url(student["photo"])
        return (
            jsonify(
                {
                    "ok": False,
                    "message": f"{student['nombres']} {student['apellidos']} ya registró su {tipo_texto} el día de hoy.",
                    "when": ts,
                    "tipo": tipo,
                    "student": {
                        "nombres": student["nombres"],
                        "apellidos": student["apellidos"],
                        "dni": student["dni"],
                        "photo_url": photo_url,
                    },
                }
            ),
            400,
        )

    att_id = insert_attendance(student["id"], iid, tipo)
    ts = now.strftime("%Y-%m-%d %H:%M:%S")
    photo_url = None
    if student["photo"]:
        photo_url = url_for("static", filename=f"uploads/{student['photo']}")
    tipo_texto = "Entrada" if tipo == "ENTRADA" else "Salida"
    return jsonify(
        {
            "ok": True,
            "message": f"{tipo_texto} registrada: {student['nombres']} {student['apellidos']} - {ts}",
            "attendance_id": att_id,
            "when": ts,
            "tipo": tipo,
            "student": {
                "nombres": student["nombres"],
                "apellidos": student["apellidos"],
                "dni": student["dni"],
                "photo_url": photo_url,
            },
        }
    )


@app.route("/students", methods=["GET", "POST"])
@tenant_required
def students():
    iid = current_institution_id()
    if request.method == "POST":
        nombres = request.form.get("nombres", "").strip()
        apellidos = request.form.get("apellidos", "").strip()
        dni = request.form.get("dni", "").strip()

        if not (nombres and apellidos and dni):
            flash("Todos los campos son obligatorios", "error")
            return redirect(url_for("students"))

        photo_filename = None
        if "photo" in request.files:
            file = request.files["photo"]
            if file and file.filename and allowed_file(file.filename):
                unique_filename = photo_name_for_save(iid, dni, file.filename)
                result = storage.upload(file, unique_filename, file.content_type or "image/jpeg")
                if result:
                    photo_filename = unique_filename

        try:
            with dbx.get_connection() as conn:
                cur = conn.cursor()
                dbx.execute(
                    cur,
                    """
                    INSERT INTO students (institution_id, nombres, apellidos, dni, photo)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (iid, nombres, apellidos, dni, photo_filename),
                )
                dbx.commit(conn)
            flash("Estudiante registrado correctamente", "success")
        except Exception as e:
            if dbx.is_unique_violation(e):
                flash("El DNI ya existe en esta institución", "error")
            else:
                flash("Error al registrar estudiante", "error")
                app.logger.exception(e)
        return redirect(url_for("students"))

    q = request.args.get("q") or ""
    with dbx.get_connection() as conn:
        cur = conn.cursor()
        if q.strip():
            like = f"%{q}%"
            dbx.execute(
                cur,
                """
                SELECT * FROM students
                WHERE institution_id = %s AND (nombres LIKE %s OR apellidos LIKE %s OR dni LIKE %s)
                ORDER BY id DESC
                """,
                (iid, like, like, like),
            )
        else:
            dbx.execute(
                cur,
                "SELECT * FROM students WHERE institution_id = %s ORDER BY id DESC",
                (iid,),
            )
        student_rows = dbx.fetchall(cur)

    return render_template("students.html", students=student_rows, q=q)


def _get_student_in_institution(sid: int, iid: int) -> Optional[Any]:
    with dbx.get_connection() as conn:
        cur = conn.cursor()
        dbx.execute(
            cur,
            "SELECT * FROM students WHERE id = %s AND institution_id = %s",
            (sid, iid),
        )
        return dbx.fetchone(cur)


@app.route("/students/<int:sid>/edit", methods=["GET", "POST"])
@tenant_required
def edit_student(sid: int):
    iid = current_institution_id()
    student = _get_student_in_institution(sid, iid)
    if not student:
        flash("Estudiante no encontrado", "error")
        return redirect(url_for("students"))

    if request.method == "POST":
        nombres = request.form.get("nombres", "").strip()
        apellidos = request.form.get("apellidos", "").strip()
        dni = request.form.get("dni", "").strip()

        if not (nombres and apellidos and dni):
            flash("Todos los campos son obligatorios", "error")
            return redirect(url_for("edit_student", sid=sid))

        photo_filename = student["photo"]
        if "photo" in request.files:
            file = request.files["photo"]
            if file and file.filename and allowed_file(file.filename):
                if student["photo"]:
                    storage.delete(student["photo"])
                unique_filename = photo_name_for_save(iid, dni, file.filename)
                result = storage.upload(file, unique_filename, file.content_type or "image/jpeg")
                if result:
                    photo_filename = unique_filename

        try:
            with dbx.get_connection() as conn:
                cur = conn.cursor()
                dbx.execute(
                    cur,
                    """
                    UPDATE students SET nombres = %s, apellidos = %s, dni = %s, photo = %s
                    WHERE id = %s AND institution_id = %s
                    """,
                    (nombres, apellidos, dni, photo_filename, sid, iid),
                )
                dbx.commit(conn)
            flash("Estudiante actualizado correctamente", "success")
            return redirect(url_for("students"))
        except Exception as e:
            if dbx.is_unique_violation(e):
                flash("El DNI ya existe en otro registro de esta institución", "error")
            else:
                flash("Error al actualizar", "error")
                app.logger.exception(e)
            return redirect(url_for("edit_student", sid=sid))

    return render_template("edit_student.html", student=student)


@app.post("/students/<int:sid>/delete")
@tenant_required
def delete_student(sid: int):
    iid = current_institution_id()
    st = _get_student_in_institution(sid, iid)
    if not st:
        flash("Estudiante no encontrado", "error")
        return redirect(url_for("students"))

    if st["photo"]:
        storage.delete(st["photo"])

    with dbx.get_connection() as conn:
        cur = conn.cursor()
        dbx.execute(cur, "DELETE FROM attendance WHERE student_id = %s AND institution_id = %s", (sid, iid))
        dbx.execute(cur, "DELETE FROM students WHERE id = %s AND institution_id = %s", (sid, iid))
        dbx.commit(conn)
    flash("Estudiante y sus asistencias eliminados.", "success")
    return redirect(url_for("students"))


@app.route("/students/<int:sid>/qr")
@tenant_required
def student_qr(sid: int):
    iid = current_institution_id()
    student = _get_student_in_institution(sid, iid)
    if not student:
        flash("Estudiante no encontrado", "error")
        return redirect(url_for("students"))
    return render_template("student_qr.html", student=student)


@app.route("/students/<int:sid>/qr/image")
@tenant_required
def student_qr_image(sid: int):
    iid = current_institution_id()
    try:
        with dbx.get_connection() as conn:
            cur = conn.cursor()
            dbx.execute(
                cur,
                "SELECT dni, nombres, apellidos FROM students WHERE id = %s AND institution_id = %s",
                (sid, iid),
            )
            student = dbx.fetchone(cur)

        if not student:
            app.logger.error("Estudiante con id %s no encontrado en inst %s", sid, iid)
            return "Estudiante no encontrado", 404

        qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=4)
        qr.add_data(student["dni"])
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        img_io = BytesIO()
        img.save(img_io, "PNG")
        img_io.seek(0)

        return Response(
            img_io.getvalue(),
            mimetype="image/png",
            headers={
                "Content-Disposition": f'inline; filename=qr_{student["dni"]}.png',
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Pragma": "no-cache",
                "Expires": "0",
            },
        )
    except Exception as e:
        app.logger.error("Error generando QR para estudiante %s: %s", sid, str(e))
        return f"Error generando QR: {str(e)}", 500


@app.post("/students/upload-excel")
@tenant_required
def upload_excel():
    iid = current_institution_id()
    if "excel_file" not in request.files:
        flash("No se seleccionó ningún archivo", "error")
        return redirect(url_for("students"))

    file = request.files["excel_file"]
    if not file or not file.filename:
        flash("No se seleccionó ningún archivo", "error")
        return redirect(url_for("students"))

    if not file.filename.endswith((".xlsx", ".xls")):
        flash("El archivo debe ser Excel (.xlsx o .xls)", "error")
        return redirect(url_for("students"))

    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        added = 0
        skipped = 0
        errors = []

        with dbx.get_connection() as conn:
            cur = conn.cursor()
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) < 3:
                    continue
                nombres, apellidos, dni = row[0], row[1], row[2]
                if not all([nombres, apellidos, dni]):
                    errors.append(f"Fila {row_idx}: Faltan datos requeridos")
                    skipped += 1
                    continue
                dni = str(dni).strip()
                nombres = str(nombres).strip()
                apellidos = str(apellidos).strip()
                try:
                    dbx.execute(
                        cur,
                        "INSERT INTO students (institution_id, nombres, apellidos, dni) VALUES (%s, %s, %s, %s)",
                        (iid, nombres, apellidos, dni),
                    )
                    added += 1
                except Exception as e:
                    if dbx.is_unique_violation(e):
                        errors.append(f"Fila {row_idx}: DNI {dni} ya existe")
                        skipped += 1
                    else:
                        raise
            dbx.commit(conn)

        message = f"Carga completada: {added} estudiantes agregados"
        if skipped > 0:
            message += f", {skipped} omitidos"
        if errors and len(errors) <= 5:
            message += ". Errores: " + "; ".join(errors)
        elif errors:
            message += f". {len(errors)} errores encontrados"
        flash(message, "success" if added > 0 else "error")

    except Exception as e:
        flash(f"Error al procesar el archivo: {str(e)}", "error")
        app.logger.exception(e)

    return redirect(url_for("students"))


@app.route("/reports", methods=["GET", "POST"])
@tenant_required
def reports():
    iid = current_institution_id()
    if request.method == "POST":
        try:
            fecha_desde = request.form.get("fecha_desde", "").strip()
            fecha_hasta = request.form.get("fecha_hasta", "").strip()

            sql = """
                SELECT a.id, a.recorded_at, a.tipo, s.nombres, s.apellidos, s.dni
                FROM attendance a
                JOIN students s ON s.id = a.student_id
                WHERE a.institution_id = %s
            """
            params: list = [iid]

            if fecha_desde:
                sql += " AND date(a.recorded_at) >= %s"
                params.append(fecha_desde)
            if fecha_hasta:
                sql += " AND date(a.recorded_at) <= %s"
                params.append(fecha_hasta)

            sql += " ORDER BY a.recorded_at DESC"

            with dbx.get_connection() as conn:
                cur = conn.cursor()
                dbx.execute(cur, sql, params)
                rows = dbx.fetchall(cur)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte de Asistencia"
            headers = ["ID", "Nombres", "Apellidos", "DNI", "Tipo", "Fecha y Hora"]
            ws.append(headers)

            from openpyxl.styles import Alignment, Font, PatternFill

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row in rows:
                try:
                    rv = row["recorded_at"]
                    if hasattr(rv, "isoformat"):
                        fecha_str = rv.isoformat()[:19].replace("T", " ")
                    else:
                        dt = datetime.fromisoformat(str(rv).replace("Z", "+00:00"))
                        fecha_str = dt.strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    fecha_str = str(row["recorded_at"])

                ws.append(
                    [
                        row["id"],
                        row["nombres"],
                        row["apellidos"],
                        row["dni"],
                        row["tipo"],
                        fecha_str,
                    ]
                )

            for col, w in zip("ABCDEF", [8, 20, 20, 12, 12, 20]):
                ws.column_dimensions[col].width = w

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            filename = f"reporte_asistencia_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            return Response(
                output.getvalue(),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename={filename}",
                    "Cache-Control": "no-cache, no-store, must-revalidate",
                    "Pragma": "no-cache",
                    "Expires": "0",
                },
            )

        except Exception as e:
            app.logger.error("Error generando reporte Excel: %s", str(e))
            flash(f"Error al generar el reporte: {str(e)}", "error")
            return redirect(url_for("reports"))

    with dbx.get_connection() as conn:
        cur = conn.cursor()
        dbx.execute(cur, "SELECT COUNT(*) as count FROM attendance WHERE institution_id = %s", (iid,))
        total_asistencias = dbx.fetchone(cur)["count"]
        dbx.execute(cur, "SELECT COUNT(*) as count FROM students WHERE institution_id = %s", (iid,))
        total_estudiantes = dbx.fetchone(cur)["count"]

    return render_template(
        "reports.html",
        total_asistencias=total_asistencias,
        total_estudiantes=total_estudiantes,
    )


@app.route("/uploads/<filename>")
@tenant_required
def uploads(filename):
    if storage._use_supabase():
        return redirect(storage.public_url(filename))
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)


@app.cli.command("seed")
@click.option("--institution-id", "institution_id", type=int, default=None, help="ID de institución (requerido en Postgres si no hay sesión web)")
def seed(institution_id):
    iid = institution_id
    if iid is None:
        with dbx.get_connection() as conn:
            cur = conn.cursor()
            dbx.execute(
                cur,
                "SELECT id FROM institutions WHERE approval_status = %s ORDER BY id LIMIT 1",
                ("approved",),
            )
            row = dbx.fetchone(cur)
            if row:
                iid = int(row["id"])
    if not iid:
        print("No hay instituciones. Cree una con la web o pase --institution-id=N")
        return
    data = [
        ("Juan Carlos", "Pérez García", "12345678"),
        ("María Fernanda", "López Rodríguez", "87654321"),
        ("Luis Alberto", "Sánchez Torres", "44556677"),
    ]
    with dbx.get_connection() as conn:
        cur = conn.cursor()
        for n, a, d in data:
            try:
                dbx.execute(
                    cur,
                    "INSERT INTO students (institution_id, nombres, apellidos, dni) VALUES (%s, %s, %s, %s)",
                    (iid, n, a, d),
                )
            except Exception as e:
                if not dbx.is_unique_violation(e):
                    raise
        dbx.commit(conn)
    print("Seed listo para la institución actual.")


@app.cli.command("bootstrap-admin")
def bootstrap_admin():
    """Crea admin + institución por defecto en Postgres (Supabase). Ejecutar una vez."""
    if not dbx.is_postgres():
        print("Este comando solo aplica con DATABASE_URL (PostgreSQL). En SQLite use el admin por defecto.")
        return
    username = os.environ.get("BOOTSTRAP_ADMIN_USER", "admin")
    password = os.environ.get("BOOTSTRAP_ADMIN_PASSWORD", "admin123")
    inst_name = os.environ.get("BOOTSTRAP_INSTITUTION_NAME", "Mi institución")
    base_slug = slugify_institution(inst_name)
    slug = base_slug
    ph = generate_password_hash(password)
    with dbx.get_connection() as conn:
        cur = conn.cursor()
        dbx.execute(cur, "SELECT id FROM users WHERE username = %s", (username,))
        u = dbx.fetchone(cur)
        if not u:
            dbx.execute(
                cur,
                "INSERT INTO users (username, password_hash, is_super_master) VALUES (%s, %s, %s)",
                (username, ph, True),
            )
            dbx.commit(conn)
            dbx.execute(cur, "SELECT id FROM users WHERE username = %s", (username,))
            u = dbx.fetchone(cur)
        uid = int(u["id"])
        dbx.execute(cur, "UPDATE users SET is_super_master = %s WHERE id = %s", (True, uid))
        dbx.commit(conn)
        for n in range(20):
            dbx.execute(cur, "SELECT id FROM institutions WHERE slug = %s", (slug,))
            if dbx.fetchone(cur):
                slug = f"{base_slug}-{n+2}"
                continue
            dbx.execute(
                cur,
                "INSERT INTO institutions (name, slug, approval_status) VALUES (%s, %s, %s)",
                (inst_name, slug, "approved"),
            )
            dbx.commit(conn)
            dbx.execute(cur, "SELECT id FROM institutions WHERE slug = %s", (slug,))
            irow = dbx.fetchone(cur)
            iid = int(irow["id"])
            dbx.execute(
                cur,
                """
                INSERT INTO institution_members (user_id, institution_id, role)
                VALUES (%s, %s, %s)
                ON CONFLICT (user_id, institution_id) DO NOTHING
                """,
                (uid, iid, "owner"),
            )
            dbx.commit(conn)
            print(f"Listo: usuario '{username}', institución '{inst_name}' (slug={slug}). Cambie la contraseña en producción.")
            return
    print("No se pudo asignar slug único a la institución.")


@app.errorhandler(404)
def page_not_found(e):
    return render_template("404.html"), 404


@app.errorhandler(500)
def internal_server_error(e):
    error_msg = str(e) if app.config.get("DEBUG") else None
    return render_template("500.html", error=error_msg), 500


@app.errorhandler(Exception)
def handle_exception(e):
    if isinstance(e, Exception) and hasattr(e, "code"):
        return e
    app.logger.error("Error no manejado: %s", e, exc_info=True)
    error_msg = str(e) if app.config.get("DEBUG") else None
    return render_template("500.html", error=error_msg), 500


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000, threaded=True, processes=1)
